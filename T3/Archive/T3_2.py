import urllib
import json
from requests.api import request
import pandas as pd
import numpy as np

# 函数功能：向高德地图发送请求，得到反馈
# 函数的返回值：完整的返回的 json 格式的数据
# 孟昊阳高德地图申请到的 Key：aad49afa17b46e85e060bbe252f25a80
def get_path(start_gps_x, start_gps_y, end_gps_x, end_gps_y, user_key):
    # URL Example: https://restapi.amap.com/v3/direction/driving?origin=116.481028,39.989643&destination=116.465302,40.004717&extensions=all&output=xml&key=<用户的key>
    if (str(start_gps_x).strip() != 'NaN' and str(start_gps_y).strip() != 'NaN' and str(end_gps_y).strip() != 'NaN' and str(end_gps_y).strip() != 'NaN'):
        url = 'https://restapi.amap.com/v3/direction/driving?origin=' + str(start_gps_x).strip() + ',' + str(start_gps_y).strip() + '&destination=' + str(end_gps_x).strip() + ',' + (end_gps_y).strip() + '&extensions=all&output=json&key=' + str(user_key).strip()
        try:
            returned_result = urllib.request.urlopen(url).read()
            returned_result = returned_result.decode()
            returned_result = str(returned_result)
            return returned_result
        except:
            return '-1'
    else:
        return '-2'

# 函数功能：处理返回的 json 格式数据，得到路程时间
# 函数的返回值：字符串类型的所需时间
def time_cost(returned_result):
    if (returned_result != '-1' and returned_result != '-2'):
        temp = json.loads(returned_result)
        processed_path_time = temp["route"]["paths"]
        duration = str(processed_path_time).split(',')[1].split(":")[1].strip()
        return str(duration)
    elif (returned_result == '-1'):
        return 'error: 网络错误!'
    elif (returned_result == '-2'):
        return 'error: 输入坐标不全!'

# 具体思路：本次处警后，如果没有警情就待着，有的话就朝新的警情开进，如果中途有了新的警情，不理会；到目的地后，开始处理，如果有多个新警情，优先赶往最近的。
def path_processing():
    return 

# 函数功能：按照起始和终末位置输出 HTML 文件显示规划路线
# 作者：胡文强
def output_webpage(start_gps_double, a, b, c, d, e, f, g, h, i, j, k, l, m, n, o, p, end_gps_double):
    output_html = open('T3_output_generated_by_Python.html','w',encoding='utf-8')

    output_html.write('<!doctype html>' + '\n')
    output_html.write('<html>' + '\n')
    output_html.write('<head>' + '\n')
    output_html.write('    <meta charset="utf-8">' + '\n')
    output_html.write('    <meta http-equiv="X-UA-Compatible" content="IE=edge">' + '\n')
    output_html.write('    <meta name="viewport" content="initial-scale=1.0, user-scalable=no, width=device-width">' + '\n')
    output_html.write('    <title>T3 Visualized</title>' + '\n')
    output_html.write('    <style type="text/css">' + '\n')
    output_html.write('    html,' + '\n')
    output_html.write('    body,' + '\n')
    output_html.write('    #container {' + '\n')
    output_html.write('      width: 100%;' + '\n')
    output_html.write('      height: 100%;}' + '\n')
    output_html.write('    </style>' + '\n')
    output_html.write('    <style type="text/css">' + '\n')
    output_html.write('        #panel {' + '\n')
    output_html.write('            position: fixed;' + '\n')
    output_html.write('            background-color: white;' + '\n')
    output_html.write('            max-height: 90%;' + '\n')
    output_html.write('            overflow-y: auto;' + '\n')
    output_html.write('            top: 10px;' + '\n')
    output_html.write('            right: 10px;' + '\n')
    output_html.write('            width: 280px;' + '\n')
    output_html.write('        }' + '\n')
    output_html.write('        #panel .amap-call {' + '\n')
    output_html.write('            background-color: #009cf9;' + '\n')
    output_html.write('            border-top-left-radius: 4px;' + '\n')
    output_html.write('   	        border-top-right-radius: 4px;' + '\n')
    output_html.write('        }' + '\n')
    output_html.write('        #panel .amap-lib-driving {' + '\n')
    output_html.write('	        border-bottom-left-radius: 4px;' + '\n')
    output_html.write('   	        border-bottom-right-radius: 4px;' + '\n')
    output_html.write('            overflow: hidden;' + '\n')
    output_html.write('        }' + '\n')
    output_html.write('    </style>' + '\n')
    output_html.write('	<script type="text/javascript">window._AMapSecurityConfig = {securityJsCode:\'d501da985e5303d4732185ad742be244\',}</script>' + '\n')
    output_html.write('    <link rel="stylesheet" href="https://a.amap.com/jsapi_demos/static/demo-center/css/demo-center.css" />' + '\n')
    output_html.write('    <script src="https://a.amap.com/jsapi_demos/static/demo-center/js/demoutils.js"></script>' + '\n')
    output_html.write('    <script type="text/javascript" src="https://webapi.amap.com/maps?v=1.4.15&key=b9af4a2d4b79762c87934da5010dd29c&plugin=AMap.Driving"></script>' + '\n')
    output_html.write('    <script type="text/javascript" src="https://cache.amap.com/lbs/static/addToolbar.js"></script>' + '\n')
    output_html.write('</head>' + '\n')
    output_html.write('<body>' + '\n')
    output_html.write('<div id="container"></div>' + '\n')
    output_html.write('<div id="panel"></div>' + '\n')
    output_html.write('<script type="text/javascript">' + '\n')
    output_html.write('    var map = new AMap.Map("container", {resizeEnable: true,center: [116.397428, 39.90923],zoom: 13});' + '\n')
    output_html.write('    var driving = new AMap.Driving({map: map,panel: "panel"}); ' + '\n')
    output_html.write('driving.search(new AMap.LngLat(\'')
    output_html.write(start_gps_double)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(end_gps_double)
    output_html.write('\'), ')
    output_html.write('{waypoints:[new AMap.LngLat(\'')
    output_html.write(a)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(b)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(c)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(d)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(e)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(f)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(g)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(h)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(i)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(j)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(k)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(l)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(m)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(n)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(o)
    output_html.write('\'), new AMap.LngLat(\'')
    output_html.write(p)
    output_html.write('\')}, function(status, result) {' + '\n')
    output_html.write('        if (status === \'complete\') {' + '\n')
    output_html.write('            log.success(\'Completed Drawing Driving Paths\')' + '\n')
    output_html.write('        } else {' + '\n')
    output_html.write('            log.error(\'Error:\' + result)' + '\n')
    output_html.write('        }' + '\n')
    output_html.write('    });' + '\n')
    output_html.write('</script>' + '\n')
    output_html.write('</body>' + '\n')
    output_html.write('</html>' + '\n')
    output_html.close()
    return

def output_webpage2(start_gps_double, a, b, end_gps_double):
    output_html = open('T3_output_generated_by_Python.html','w',encoding='utf-8')

    output_html.write('<!doctype html>' + '\n')
    output_html.write('<html>' + '\n')
    output_html.write('<head>' + '\n')
    output_html.write('    <meta charset="utf-8">' + '\n')
    output_html.write('    <meta http-equiv="X-UA-Compatible" content="IE=edge">' + '\n')
    output_html.write('    <meta name="viewport" content="initial-scale=1.0, user-scalable=no, width=device-width">' + '\n')
    output_html.write('    <title>T3 Visualized</title>' + '\n')
    output_html.write('    <style>' + '\n')
    output_html.write('    html,' + '\n')
    output_html.write('    body,' + '\n')
    output_html.write('    #container {' + '\n')
    output_html.write('      width: 100%;' + '\n')
    output_html.write('      height: 100%;}' + '\n')
    output_html.write('        #panel {' + '\n')
    output_html.write('            position: fixed;' + '\n')
    output_html.write('            background-color: white;' + '\n')
    output_html.write('            max-height: 90%;' + '\n')
    output_html.write('            overflow-y: auto;' + '\n')
    output_html.write('            top: 10px;' + '\n')
    output_html.write('            right: 10px;' + '\n')
    output_html.write('            width: 280px;' + '\n')
    output_html.write('        }' + '\n')
    output_html.write('        #panel .amap-lib-driving {' + '\n')
    output_html.write('            border-radius: 4px;' + '\n')
    output_html.write('             overflow: hidden' + '\n')
    output_html.write('        }' + '\n')
    output_html.write('    </style>' + '\n')
    output_html.write('	<script type="text/javascript">window._AMapSecurityConfig = {securityJsCode:\'d501da985e5303d4732185ad742be244\',}</script>' + '\n')
    output_html.write('    <link rel="stylesheet" href="https://a.amap.com/jsapi_demos/static/demo-center/css/demo-center.css" />' + '\n')
    output_html.write('    <script src="https://a.amap.com/jsapi_demos/static/demo-center/js/demoutils.js"></script>' + '\n')
    output_html.write('    <script type="text/javascript" src="https://webapi.amap.com/maps?v=1.4.15&key=b9af4a2d4b79762c87934da5010dd29c&plugin=AMap.Driving"></script>' + '\n')
    output_html.write('    <script type="text/javascript" src="https://cache.amap.com/lbs/static/addToolbar.js"></script>' + '\n')
    output_html.write('</head>' + '\n')
    output_html.write('<body>' + '\n')
    output_html.write('<div id="container"></div>' + '\n')
    output_html.write('<div id="panel"></div>' + '\n')
    output_html.write('<script type="text/javascript">' + '\n')
    output_html.write('    var map = new AMap.Map("container", {resizeEnable: true,center: [116.397428, 39.90923],zoom: 13});' + '\n')
    output_html.write('    var driving = new AMap.Driving({map: map,panel: "panel"}); ' + '\n')
    output_html.write('driving.search(new AMap.LngLat(')
    output_html.write(start_gps_double)
    output_html.write('), new AMap.LngLat(')
    output_html.write(end_gps_double)
    output_html.write('), ')
    output_html.write('{waypoints:[new AMap.LngLat(')
    output_html.write(a)
    output_html.write('), new AMap.LngLat(')
    output_html.write(b)
    output_html.write(')]}, function(status, result) {' + '\n')
    output_html.write('        if (status === \'complete\') {' + '\n')
    output_html.write('            log.success(\'Completed Drawing Driving Paths\')' + '\n')
    output_html.write('        } else {' + '\n')
    output_html.write('            log.error(\'Error:\' + result)' + '\n')
    output_html.write('        }' + '\n')
    output_html.write('    });' + '\n')
    output_html.write('</script>' + '\n')
    output_html.write('</body>' + '\n')
    output_html.write('</html>' + '\n')
    output_html.close()
    return
# 函数功能：将12.1日报警信息按时间排序后输出excel表格，包括报警人基本信息，结合实际
# 作者：孟昊阳
def get_12_1():
    try:
        df = pd.read_excel(r'./GPS数据结果.xlsx')
    except:
        print("Error: 输入文件 (GPS数据结果.xlsx) 错误!")
        quit()
    df_gps = df.sort_values('报警时间', ascending=True)
    df_gps = df_gps[:-1]
    data = df_gps[df_gps.报警时间.str.startswith("2020-12-01")]
    data.to_excel(r'./2020.12.1.xlsx', index=False, header=True)
    from openpyxl import load_workbook

    wb = load_workbook(r'./2020.12.1.xlsx')
    ws = wb[wb.sheetnames[0]]  # 打开第一个sheet
    ws.column_dimensions['A'].width = 31.0  # 调整列A宽
    ws.column_dimensions['B'].width = 19.0
    ws.column_dimensions['C'].width = 8.0
    ws.column_dimensions['D'].width = 15.0
    ws.column_dimensions['E'].width = 20.0
    ws.column_dimensions['F'].width = 20.0
    ws.column_dimensions['G'].width = 80.0
    wb.save(r'./2020.12.1.xlsx')
def putandout(x):
    x = x.tolist()
    x1 = x[0]
    x2 = x[1]
    x3 = x[2]
    x4 = x[3]
    x5 = x[4]
    x6 = x[5]
    x7 = x[6]
    x8 = x[7]
    x9 = x[8]
    x10 = x[9]
    x11 = x[10]
    x12 = x[11]
    x13 = x[12]
    x14 = x[13]
    x15 = x[14]
    x16 = x[15]
    x17 = x[16]
    x18 = x[17]
    output_webpage(x1,x2,x3,x4,x5,x6,x7,x8,x9,x10,x11,x12,x13,x14,x15,x16,x17,x18)

def putandout2(x):
    x = x.tolist()
    x1 = x[0]
    x2 = x[1]
    x3 = x[2]
    x4 = x[3]
    output_webpage2(x1,x2,x3,x4)

if __name__ == "__main__":
    get_12_1()
    try:
        GPS_double = pd.read_excel(r'./2020.12.1.xlsx')
    except:
        print("Error: 输入文件 (2020.12.1.xlsx) 错误!")
        quit()
    GPS_double['GPS定位数对'] = GPS_double['GPS_x'].map(str) + ',' + GPS_double['GPS_y'].map(str)
    GPS_double = GPS_double.drop(index=(GPS_double.loc[(GPS_double['GPS定位数对'] == 'nan,nan')].index))
    GPS_double = GPS_double['GPS定位数对']
    # GPS_double1 = GPS_double[i:a]

    a = np.arange(0, len(GPS_double), 17)
    b = np.arange(18, len(GPS_double) + 17, 17)
    a = a.tolist()
    b = b.tolist()
    List = zip(a, b)
    GPS_double3 = pd.DataFrame()
    for s, e in List:
        GPS_double2 = GPS_double[s:e]
        GPS_double2.index = range(1, len(GPS_double2) + 1)
        GPS_double3 = GPS_double3.append(GPS_double2, ignore_index=True)
    GPS_double3.index = range(1, len(GPS_double3) + 1)
    data1 = GPS_double3.iloc[0]
    data2 = GPS_double3.iloc[1]
    data3 = GPS_double3.iloc[2]
    data4 = GPS_double3.iloc[3]
    data5 = GPS_double3.iloc[4]
    data6 = GPS_double3.iloc[5]
    data7 = GPS_double3.iloc[6]
    data7 = data7.iloc[0:4]
    #result = get_path('119.944296584982','30.097262635875','119.945421709799','30.045358805853','aad49afa17b46e85e060bbe252f25a80')
    putandout2(data7)
    #print(time_cost(result))
