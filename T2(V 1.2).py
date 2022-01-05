# 第二题题目：
# 根据标注地点从地图API获取其具体坐标

import urllib
import json
from requests.api import request
from pandas import Series, DataFrame
import pandas as pd
import time
import matplotlib.pyplot as plt
# from bs4 import BeautifulSoup
# import re
# import json
# import pycurl
# from io import BytesIO
# import requests

##################################################################################
# 函数功能：调取高德地图API获得信息
# 作者：胡文强
# 主要思路：通过获取的数据，构造URL并通过requests库向高德地图 API 访问得到返回值，再进行数据处理得到有效信息
# 具体实现：
# 1. 获取值：
#   1. returned_information_format: 返回信息的类型：json 还是 xml，我觉得json更方便分析
#   2. input_x, input_y: 输入的 GPS 的 X、Y 坐标
#   3. input_key: 用户 Key
#   4. result_radius: 查询的半径
#   5. returned_information_kind: 是返回所有的结果（all）还是单个（base）
# 2. 校验值是否为 NaN，如果是则直接返回 '-2', 表示输入数据有误
# 3. 通过 urllib 的 request 来调取API得到返回值，如果正常则返回 string 类型的 returned_result，如果有异常则返回 string 类型的 '-1'
# 4. 将返回值给别的函数（get_processed_location）处理，得到所需数据
# 备注：通过孟昊阳的高德地图账号申请到 key 为：aad49afa17b46e85e060bbe252f25a80
def get_location(returned_information_format, input_x, input_y, input_key, returned_information_kind):
    # Url example: https://restapi.amap.com/v3/geocode/regeo?output=xml&location=116.310003,39.991957&key=<用户的key>&radius=1000&extensions=all (all/base)
    if (input_x.strip() != 'NaN' and input_y.strip() != 'NaN'):
        url = 'https://restapi.amap.com/v3/geocode/regeo?output=' + str(
            returned_information_format).strip() + 'xml&location=' + str(input_x).strip() + ',' + str(
            input_y).strip() + '&key=' + str(input_key).strip() + '&radius=1000' + '&extensions=' + str(
            returned_information_kind).strip()
        try:
            returned_result = urllib.request.urlopen(url).read()
            returned_result = returned_result.decode()
            returned_result = str(returned_result)
            return returned_result
        except:
            return '-1'
    else:
        return '-2'
    #############################################################
    # try:
    #     # returned_result = requests.get(url, timeout=300)
    #     returned_result = urllib3.urlopen(url)
    #     return returned_result
    # except:
    #     return -1
    # with urllib.request.urlopen(url) as response:
    #     returned_result = response.read()
    #     returned_result.encode(encoding='utf-8',errors='strict')
    #     return returned_result
    #############################################################
    #############################################################
    # 以下可行但是会乱码
    # try:
    #     returned_result = pycurl.Curl()
    #     returned_result.setopt(pycurl.URL, url) 
    #     returned_result.setopt(pycurl.SSL_VERIFYPEER, False)
    #     returned_result.perform()
    #     returned_result.encode(encoding='utf-8',errors='strict')
    #     #returned_result.encode('gb2312') 
    #     return returned_result
    # except:
    #     return -1
    #############################################################


##################################################################################

##################################################################################
# 函数功能：处理通过地图API获取到的json信息得到地址输出
# 作者：胡文强
# 主要思路：先进行校验，若收录则对信息进行处理，若无法得到任何结果则返回 “未收录”
# 具体实现：
# 1. 进行校验，若 return_result 不为 -1 （即网络正常，可访问）则开始处理：
#   1. 将 returned_result 转为json数据
#   2. 从中提取出 regeocode 中的 formatted_address 项，如果失败返回 “error: 坐标有误！”
# 2. 若为 '-1' 则返回结果 “error: 网络错误！”
# 3. 若为 '-2' 则返回结果 “error: 输入坐标不全!”
def get_processed_location(returned_result):
    if (returned_result != '-1' and returned_result != '-2'):
        try:
            temp = json.loads(returned_result)
            processed_location = temp["regeocode"]["formatted_address"]
            return processed_location
        except:
            return 'error: 坐标有误!'
    elif (returned_result == '-1'):
        return 'error: 网络错误!'
    elif (returned_result == '-2'):
        return 'error: 输入坐标不全!'
##################################################################################

##################################################################################
# 函数功能：读取表格中的数据并处理，完成后进行输出
# 作者：孟昊阳，胡文强
# 主要思路：读入表格中的GPS列信息，利用上述函数进行处理，整合之后输出到Excel中
# 具体实现：如代码中注释所述
if __name__ == "__main__":
    # 计时
    T1 = time.time()

    # result = get_location('json', '119.944296584982', '30.097262635875', 'aad49afa17b46e85e060bbe252f25a80', 'base')
    # print(get_processed_location(result))
    # 引入考试数据
    df = pd.read_excel("./考试数据.xls")  # 被胡文强改为了相对路径
    # 对GPS两列进行数据提取
    GPS_get_x = df['GPS_X'].to_string(header=False, index=False).split('\n')
    GPS_get_x = GPS_get_x[1:]

    GPS_get_y = df['GPS_Y'].to_string(header=False, index=False).split('\n')
    GPS_get_y = GPS_get_y[1:]
    # AGEADDRESS_list = pd.DataFrame()#建造一个空列表
    output_list = pd.DataFrame()
    # 将GPS提取并处理得到具体地址
    step_counter = 1
    for GPS_x, GPS_y in zip(GPS_get_x, GPS_get_y):
        GPS = get_processed_location(get_location('json', GPS_x, GPS_y, 'aad49afa17b46e85e060bbe252f25a80', 'base'))
        output_list = output_list.append(pd.DataFrame({'GPS地址': [GPS]}), ignore_index=True)
        print(step_counter,": ", GPS)
        step_counter = step_counter +1


    data3 =  [{'GPS地址': 'GPS地址'}]
    df_blank3 = pd.DataFrame(data3)
    output_list = df_blank3.append(output_list, ignore_index=True)

    #############################################################
    df_base = df['Unnamed: 7']
    df_base2 = df['Unnamed: 8']
    df_base = df_base[1:]
    df_base2 = df_base2[1:]
    df_date = df['报警时间']
    df_date = df_date[1:]

    df_base3 = pd.concat([df_base, df_base2], axis=1)

    df_phone = df['Unnamed: 3']
    df_base3 = pd.concat([df_base3, df_phone], axis=1)
    df_base3 = df_base3[1:]
    df_base3 = pd.concat([df_date, df_base3], axis=1)
    df_base3 = df_base3.rename(columns={'报警时间': '报警时间', 'Unnamed: 7': '身份证号码', 'Unnamed: 8': '姓名', 'Unnamed: 3': '联系人电话'})

    data2 = [{'报警时间': '报警时间', '身份证号码': '身份证号码', '姓名': '姓名', '联系人电话': '联系人电话'}]
    df_blank2 = pd.DataFrame(data2)
    df_base3 = df_blank2.append(df_base3, ignore_index=True)
    output_list = pd.concat([df_base3, output_list], axis=1)

    #############################################################

    output_list.to_excel(r'./GPS数据结果.xlsx', index=False, header=True)
    #############################################################
    ###设置excel表格格式
    from openpyxl import load_workbook

    wb = load_workbook(r'./GPS数据结果.xlsx')
    ws = wb[wb.sheetnames[0]]  # 打开第一个sheet
    ws.column_dimensions['A'].width = 31.0  # 调整列A宽
    ws.column_dimensions['B'].width = 19.0
    ws.column_dimensions['C'].width = 8.0
    ws.column_dimensions['D'].width = 15.0
    ws.column_dimensions['E'].width = 80.0
    wb.save(r'./GPS数据结果.xlsx')
    # 计时
    T2 = time.time()

    print('程序运行时间:%s秒' % (T2 - T1))
##################################################################################
