import pandas as pd
import numpy as np

df = pd.read_excel(r'./GPS数据结果.xlsx')
df_gps = df.sort_values('报警时间', ascending=False)
df_gps = df_gps[:-1]
data = df_gps[df_gps.报警时间.str.startswith("2020-12-01")]
data.to_excel(r'./2020.12.1.xlsx', index = False, header=True)
from openpyxl import load_workbook

wb = load_workbook(r'./2020.12.1.xlsx')
ws = wb[wb.sheetnames[0]]  # 打开第一个sheet
ws.column_dimensions['A'].width = 31.0  # 调整列A宽
ws.column_dimensions['B'].width = 19.0
ws.column_dimensions['C'].width = 8.0
ws.column_dimensions['D'].width = 15.0
ws.column_dimensions['E'].width = 20.0
ws.column_dimensions['F'].width = 20.0
ws.column_dimensions['G'].width = 80.0
wb.save(r'./2020.12.1.xlsx')