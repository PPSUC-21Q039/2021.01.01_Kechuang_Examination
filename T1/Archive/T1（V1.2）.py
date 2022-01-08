# 第一题题目：
# 根据所给出的事故双方的当事人的身份证号，判断其身份证号是否合法，并对身份证号、电话号码进行解析（ id_validator，phonenumber）获取其年龄、户籍、目前通讯地址（电话号码所在地）。通过统计身份证号，以及所给人员的信息，对事故多发人员的户籍地、年龄、居住社区，并进行可视化，对高发事故人员画像。
# 
# 注释要求：函数的第一行描述函数功能，第二行写作者，第三行描述具体思路，第四行分点完成具体实现

import phone 
from id_validator import validator
import time
import openpyxl
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
from pandas import Series, DataFrame
from concurrent.futures import ProcessPoolExecutor, ThreadPoolExecutor, wait, ALL_COMPLETED



list1 = [134, 135, 136, 137, 138, 139, 147, 150, 151, 152, 157, 158, 159, 178, 182, 183, 184, 187, 188,
         198, 130, 131, 132, 155, 156, 185, 186, 145, 146, 166, 175, 176, 133, 153, 177, 180, 181, 189, 173, 199, 191]
list2 = [162, 165, 167, 170, 171]

# 函数功能：电话号码归属地查询（即目前通讯地址）
# 作者：胡文强,孟昊阳（加了校检）
# 主要思路：通过 phone 库来获得电话号码信息并且提取、格式化输出为所需归属地信息
# 具体实现：
# 1. 函数获取 input_phone_number, 并将使用 phone.Phone().find() 获取具体信息，将返回值存储在 returned_phone_info 中
#    样式：{'phone': '180xxxxxxxx', 'province': '四川', 'city': 'xx', 'zip_code': 'xxxxxx', 'area_code': 'xxxx', 'phone_type': '电信'}
# 2. 之后将获取格式化的信息转为 string 类型，避免类型错误
# 3. 之后用 split 以 ',' 将返回值分隔成多个部分，取第 2 部分 (从零开始计算，故取 1): 'province': '四川', 与第 3 部分：'city': 'xx',
# 4. 分别用 split 处理得到参数的具体值并去除单引号（replace()）和空格（strip()），然后相加得到最终需要的归属地信息
def get_phone_information_address(input_phone_number): #社区
    if (input_phone_number.strip() != 'NaN'):
        if len(input_phone_number) == 11 and (int(input_phone_number[:3]) in list1):
            returned_phone_info = phone.Phone().find(input_phone_number)
            returned_phone_info_address = str(returned_phone_info).split(',')[1].split(':')[1].replace("'","").strip() + str(returned_phone_info).split(',')[2].split(':')[1].replace("'","").strip()
            return returned_phone_info_address
        elif len(input_phone_number) == 11 and (int(input_phone_number[:3]) in list2):
            return 'error: 虚拟号码!'
        else:
            return 'error: 非法号码!'
    else:
        return 'error: 无输入号码!'

# 函数功能：根据身份证号查询年龄
# 作者：胡文强
# 主要思路：通过 id_validator 库获取身份证号码信息并且提取、格式化为所需的年龄信息（暂时强制转换为 str 类型的数值，按需修改）
# 具体实现：
# 1. 首先使用 validator.is_valid() 验证身份证号是否合法（valid)，不合法的话直接返回 -1 异常值
# 2. 通过 validator.get_info() 得到身份证号具体信息，将返回值存储在 returned_id_info 中
#    样式：{'address_code': 'xxxxxx', 'abandoned': 0, 'address': '四川省xxx', 'address_tree': ['四川省', 'xxx', 'xx'], 'age': 19, 'birthday_code': '2003-08-08', 'constellation': '狮子座', 'chinese_zodiac': '未羊', 'sex': 1, 'length': 18, 'check_bit': '0'}
# 3. 之后将 returned_id_info 强制转换为 string 型变量，通过 split() 使用','对字符串进行分隔，取第7部分（由0开始计算，故值为6）
# 4. 然后将提出出来的值再次 split()，取第二部分，并通过 strip() 去除空格，之后返回 string 类型的年龄值
def get_id_information_age(input_id_number): #年龄
    if validator.is_valid(str(input_id_number)) and input_id_number.strip() != 'NaN':
        returned_id_info = validator.get_info(str(input_id_number))
        returned_id_info_age = str(returned_id_info).split(',')[6].split(':')[1].strip() 
        return str(returned_id_info_age)
    elif input_id_number.strip == 'NaN':
        return ''
    else:
        return 'error: 身份证号有误，无法处理年龄!'

# 函数功能：根据身份证号查询户籍地址
# 作者：胡文强
# 主要思路：通过 id_validator 库获取身份证号码信息并且提取、格式化为所需的户籍信息（暂时强制转换为 str 类型的数值，按需修改）
# 具体实现：
# 1. 首先使用 validator.is_valid() 验证身份证号是否合法（valid)，不合法的话直接返回 -1 异常值
# 2. 通过 validator.get_info() 得到身份证号具体信息，将返回值存储在 returned_id_info 中
#    样式：{'address_code': 'xxxxxx', 'abandoned': 0, 'address': '四川省xxx', 'address_tree': ['四川省', 'xxx', 'xx'], 'age': 19, 'birthday_code': '2003-08-08', 'constellation': '狮子座', 'chinese_zodiac': '未羊', 'sex': 1, 'length': 18, 'check_bit': '0'}
# 3. 之后将 returned_id_info 强制转换为 string 型变量，通过 split() 使用','对字符串进行分隔，取第3部分（由0开始计算，故值为2）
# 4. 然后将提出出来的值再次 split()，取第二部分，由于数值存在单引号，故使用 replace() 将其替换为 NULL，并通过 strip() 去除空格，之后返回 string 类型的户籍地址
def get_id_information_address(input_id_number):
    if validator.is_valid(str(input_id_number)) and input_id_number.strip() != 'NaN':
        returned_id_info = validator.get_info(str(input_id_number))
        returned_id_info_address = str(returned_id_info).split(',')[2].split(':')[1].replace("'","").strip() 
        return returned_id_info_address
    elif input_id_number.strip == 'NaN':
        return ''
    else:
        return 'error: 身份证号有误，无法处理户籍地址!'

# 代码功能：将姓名，身份证，住址，年龄，社区等信息放在excel表格内   ——V 0.5 Demo
# 作者：孟昊阳
# 主要思路：
# 具体实现:
def main_function():
    start_time = time.time()
    # 引入考试数据
    df = pd.read_excel("./考试数据.xls")
    # 对身份证一列进行数据提取
    ID_get = df['Unnamed: 7'].to_string(header=False, index=False).split('\n')
    ID_get = ID_get[1:]
    # 下面五行是为结果建立基本已有信息：姓名+身份证号
    df_base = df['Unnamed: 7']
    df_base2 = df['Unnamed: 8']
    df_base = df_base[1:]
    df_base2 = df_base2[1:]
    # 将姓名和身份证整合到一个Datafarme（df_base3）
    df_base3 = pd.concat([df_base, df_base2], axis=1)
    ### 后期加了一个电话号码进去
    df_phone = df['Unnamed: 3']
    df_base3 = pd.concat([df_base3, df_phone], axis=1)
    df_base3 = df_base3[1:]
    df_base3 = df_base3.rename(columns={'Unnamed: 7':'身份证号码', 'Unnamed: 8':'姓名', 'Unnamed: 3':'联系人电话'})
    ##### 后期加入了表头更改以及新Darafarme的整理
    data2 = [{'身份证号码': '身份证号码', '姓名': '姓名', '联系人电话': '联系人电话'}]
    df_blank2 = pd.DataFrame(data2)
    df_base3 = df_blank2.append(df_base3, ignore_index=True)

    # 建立年龄和地址的空Datafarme框架
    AGEADDRESS_list = pd.DataFrame()
    # 利用for循环历遍每一个信息，并用函数处理分析
    for ID in (ID_get):
        AGE = get_id_information_age(ID)
        ADDRESS = get_id_information_address(ID)
        AGEADDRESS_list = AGEADDRESS_list.append(pd.DataFrame({'年龄': [AGE], '地址': [ADDRESS]}), ignore_index=True)
    # 将得到年龄地址Datafarme和之前的已有信息Datafarme进行合并

    # 提取电话号信息
    Phone_get = df['Unnamed: 3'].to_string(header=False, index=False).split('\n')
    Phone_get = Phone_get[1:]
    # 建立空Datafarme来装电话信息
    Phone_list = pd.DataFrame()
    # 同理利用for循环历遍电话号并进行处理
    for Phone_number in (Phone_get):
        i = Phone_number.replace(" ","")
        Number = get_phone_information_address(i)
        Phone_list = Phone_list.append(pd.DataFrame({'归属地': [Number]}), ignore_index=True)
    
    # 对Datafarme进行进一步整理
    data = [{'年龄': '年龄', '地址': '地址', '归属地': '归属地'}]
    df_blank = pd.DataFrame(data)
    # 将多个Datafarme进行合并处理
    AGEADDRESSPhone_list = pd.concat([AGEADDRESS_list, Phone_list], axis=1)
    AGEADDRESSPhone_list = df_blank.append(AGEADDRESSPhone_list, ignore_index=True)
    AGEADDRESSPhone_list = pd.concat([df_base3 ,AGEADDRESSPhone_list], axis=1)
    AGEADDRESSPhone_list.to_excel(r'./数据结果.xlsx', index = False, header=True)

    end_time = time.time()
    print ("Used Time: ", end_time - start_time,'s')
    # 后期维护升级内容： 1.提高运行速度，方向：利用pool池子多线程，或者，利用迭代（暂时还没查到）,或者用Numba暴力加速，但这个库只支持函数，要想办法把内容变成函数
    #                  2.加入计时器 (已完成)
    #                  3.利用numpy和matlop进行可视化（同样要考虑时间优化问题），或者直接在excel上进行图表展示
    #                  4.大部分可以简化，之后有时间再弄吧

if __name__ == "__main__":
    main_function()
    from openpyxl import load_workbook

    wb = load_workbook(r'./数据结果.xlsx')
    ws = wb[wb.sheetnames[0]]  # 打开第一个sheet
    ws.column_dimensions['A'].width = 19.0  # 调整列A宽
    ws.column_dimensions['B'].width = 8.0
    ws.column_dimensions['C'].width = 20.0
    ws.column_dimensions['D'].width = 7.0
    ws.column_dimensions['E'].width = 40.0
    ws.column_dimensions['F'].width = 13.0
    wb.save(r'./数据结果.xlsx')
