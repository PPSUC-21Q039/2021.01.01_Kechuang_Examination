# 第二题题目：
# 根据标注地点从地图API获取其具体坐标

# 程序输入输出介绍：
# 1. 程序输入：程序输入文件为本次发放的考试数据，文件名为：考试数据.xls
# 2. 程序输出：
#   1. 程序终端输出：
#       1. 获取并处理后的地址信息
#       2. 本程序运行所用时间
#   2. 程序文件输出：
#       1. GPS数据结果.xlsx，为程序处理后得到的地址信息配合其他基本信息组合得到的输出
#       2. GPS数据对.xlsx，为程序处理后用于可视化分析事故多发区域的数据集，该数据集不做作业提交只用，仅用于提交至高德数据平台进行可视化分析。分析结果参见本目录下的：高德可视化.html 或访问： https://maplab.amap.com/share/mapv/ece53f4b48cc4d6c24faaedd2948f1fe


import urllib
import json
from requests.api import request
import pandas as pd
import time
from openpyxl import load_workbook
##################################################################################
# 函数功能：调取高德地图API获得信息
# 作者：胡文强
# 主要思路：通过获取的数据，构造URL并通过requests库向高德地图 API 访问得到返回值，再进行数据处理得到有效信息
# 具体实现：
# 1. 获取值：
#   1. returned_information_format: 返回信息的类型：json 还是 xml，我觉得json更方便分析
#   2. input_x, input_y: 输入的 GPS 的 X、Y 坐标
#   3. input_key: 用户 Key
#   4. result_radius: 查询的半径
#   5. returned_information_kind: 是返回所有的结果（all）还是单个（base）
# 2. 校验值是否为 NaN，如果是则直接返回 '-2', 表示输入数据有误
# 3. 通过 urllib 的 request 来调取API得到返回值，如果正常则返回 string 类型的 returned_result，如果有异常则返回 string 类型的 '-1'
# 4. 将返回值给别的函数（get_processed_location）处理，得到所需数据
# 备注：通过孟昊阳的高德地图账号申请到 key 为：aad49afa17b46e85e060bbe252f25a80
def get_location(returned_information_format, input_x, input_y, input_key, returned_information_kind):
    # Url example: https://restapi.amap.com/v3/geocode/regeo?output=xml&location=116.310003,39.991957&key=<用户的key>&radius=1000&extensions=all (all/base)
    if (input_x.strip() != 'NaN' and input_y.strip() != 'NaN'):
        url = 'https://restapi.amap.com/v3/geocode/regeo?output=' + str(
            returned_information_format).strip() + 'xml&location=' + str(input_x).strip() + ',' + str(
            input_y).strip() + '&key=' + str(input_key).strip() + '&radius=1000' + '&extensions=' + str(
            returned_information_kind).strip()
        try:
            returned_result = urllib.request.urlopen(url).read()
            returned_result = returned_result.decode()
            returned_result = str(returned_result)
            return returned_result
        except:
            return '-1'
    else:
        return '-2'
##################################################################################

##################################################################################
# 函数功能：处理通过地图API获取到的json信息得到地址输出
# 作者：胡文强
# 主要思路：先进行校验，若收录则对信息进行处理，若无法得到任何结果则返回 “未收录”
# 具体实现：
# 1. 进行校验，若 return_result 不为 -1 （即网络正常，可访问）则开始处理：
#   1. 将 returned_result 转为json数据
#   2. 从中提取出 regeocode 中的 formatted_address 项，如果失败返回 “error: 坐标有误！”
# 2. 若为 '-1' 则返回结果 “error: 网络错误！”
# 3. 若为 '-2' 则返回结果 “error: 输入坐标不全!”
def get_processed_location(returned_result):
    if (returned_result != '-1' and returned_result != '-2'):
        try:
            temp = json.loads(returned_result)
            processed_location = temp["regeocode"]["formatted_address"]
            return processed_location
        except:
            return 'error: 坐标有误!'
    elif (returned_result == '-1'):
        return 'error: 网络错误!'
    elif (returned_result == '-2'):
        return 'error: 输入坐标不全!'
##################################################################################

##################################################################################
def set_format():
    try:
        wb = load_workbook(r'./GPS数据结果.xlsx')
    except:
        print("Error: 文件 (GPS数据结果.xlsx) 被占用!")
        quit()
    ws = wb[wb.sheetnames[0]]  # 打开第一个sheet
    ws.column_dimensions['A'].width = 31.0  # 调整列A宽
    ws.column_dimensions['B'].width = 19.0
    ws.column_dimensions['C'].width = 8.0
    ws.column_dimensions['D'].width = 15.0
    ws.column_dimensions['E'].width = 20.0
    ws.column_dimensions['F'].width = 20.0
    ws.column_dimensions['G'].width = 80.0
    try:
        wb.save(r'./GPS数据结果.xlsx')
    except:
        print("Error: 文件 (GPS数据结果.xlsx) 被占用!")
        quit()
    return
##################################################################################

##################################################################################
# 函数功能：读取表格中的数据并处理，完成后进行输出
# 作者：孟昊阳，胡文强
# 主要思路：读入表格中的GPS列信息，利用上述函数进行处理，之后输出到Excel中
# 具体实现：如代码中注释所述
def GPS_GET():
    try:
        df = pd.read_excel("./考试数据.xls")  # 被胡文强改为了相对路径
    except:
        print("Error: 输入文件 (考试数据.xls) 错误!")
        quit()
    # 对GPS两列进行数据提取
    GPS_get_x = df['GPS_X'].to_string(header=False, index=False).split('\n')
    GPS_get_x = GPS_get_x[1:]

    GPS_get_y = df['GPS_Y'].to_string(header=False, index=False).split('\n')
    GPS_get_y = GPS_get_y[1:]
    # AGEADDRESS_list = pd.DataFrame()#建造一个空列表
    output_list = pd.DataFrame()
    # 将GPS提取并处理得到具体地址
    step_counter = 1
    for GPS_x, GPS_y in zip(GPS_get_x, GPS_get_y):
        GPS = get_processed_location(get_location('json', GPS_x, GPS_y, 'aad49afa17b46e85e060bbe252f25a80', 'base'))
        output_list = output_list.append(pd.DataFrame({'GPS地址': [GPS]}), ignore_index=True)
        print(step_counter, ": ", GPS)
        step_counter = step_counter + 1

    data3 = [{'GPS地址': 'GPS地址'}]
    df_blank3 = pd.DataFrame(data3)
    output_list = df_blank3.append(output_list, ignore_index=True)

    ################################################
    df_base = df['Unnamed: 7']
    df_base2 = df['Unnamed: 8']
    df_base = df_base[1:]
    df_base2 = df_base2[1:]
    df_date = df['报警时间']
    df_date = df_date[1:]
    GPS_X = df['GPS_X']
    GPS_X = GPS_X[1:]
    GPS_Y = df['GPS_Y']
    GPS_Y = GPS_Y[1:]
    df_GPS = pd.concat([GPS_X, GPS_Y], axis=1)

    df_base3 = pd.concat([df_base, df_base2], axis=1)

    df_phone = df['Unnamed: 3']
    df_base3 = pd.concat([df_base3, df_phone], axis=1)
    df_base3 = df_base3[1:]
    df_base3 = pd.concat([df_date, df_base3], axis=1)
    df_base3 = pd.concat([df_base3, df_GPS], axis=1)
    df_base3 = df_base3.rename(
        columns={'报警时间': '报警时间', 'Unnamed: 7': '身份证号码', 'Unnamed: 8': '姓名', 'Unnamed: 3': '联系人电话', 'GPS_X': 'GPS_x',
                 'GPS_Y': 'GPS_y'})

    data2 = [{'报警时间': '报警时间', '身份证号码': '身份证号码', '姓名': '姓名', '联系人电话': '联系人电话', 'GPS_x': 'GPS_x', 'GPS_y': 'GPS_y'}]
    df_blank2 = pd.DataFrame(data2)
    df_base3 = df_blank2.append(df_base3, ignore_index=True)
    output_list = pd.concat([df_base3, output_list], axis=1)

    try:
        output_list.to_excel(r'./GPS数据结果.xlsx', index=False, header=True)
    except:
        print("Error: 文件 (GPS数据结果.xlsx) 被占用!")
        quit()
##################################################################################

##################################################################################
def GPS_double_get():
    try:
        GPS_double = pd.read_excel("./考试数据.xls")  # 被胡文强改为了相对路径
    except:
        print("Error: 输入文件 (考试数据.xls) 错误!")
        quit()
    GPS_double['GPS定位数对'] = GPS_double['GPS_x'].map(str) + ',' + GPS_double['GPS_y'].map(str)
    GPS_double = GPS_double['GPS定位数对']
    GPS_double = GPS_double[1:]
    GPS_double.to_excel(r'./GPS数据对.xlsx', index=False, header=True)
    wb = load_workbook(r'./GPS数据对.xlsx')
    ws = wb[wb.sheetnames[0]]  # 打开第一个sheet
    ws.column_dimensions['A'].width = 35.0  # 调整列A宽
    try:
        wb.save(r'./GPS数据对.xlsx')
    except:
        print("Error: 文件 (GPS数据对.xlsx) 被占用!")
        quit()
##################################################################################

##################################################################################
if __name__ == "__main__":
    T1 = time.time() # 计时
    GPS_GET()
    GPS_double_get()
    set_format() # 设置excel表格格式
    T2 = time.time() # 计时
    print('程序运行时间:%s秒' % (T2 - T1))
##################################################################################
